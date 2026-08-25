from __future__ import annotations

import math
import os
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd
from openpyxl import Workbook
from presidio_anonymizer import AnonymizerEngine
from presidio_anonymizer.entities import OperatorConfig

from beyond_recognizers import (
    create_beyond_analyzer,
    residual_scan,
    scan_missed_proper_nouns,
    PERSON_FALSE_POSITIVE_ALLOW_LIST,
)


def _worker_process_count() -> int:
    """
    Worker processes for Presidio's batch analyzer, sized to the machine
    actually running the tool rather than a value tuned on one dev box.
    Leaves a core free for the UI thread/OS so the app stays responsive
    while a run is in progress; caps at 8 since NLP throughput gains
    flatten out well before most machines' core counts.
    """
    cpu_count = os.cpu_count() or 1
    return max(1, min(cpu_count - 1, 8))


def _analysis_batch_size() -> int:
    """
    Docs per internal spaCy nlp.pipe() batch. Scales mildly with core count
    since a machine with more workers can usefully keep more docs in flight,
    bounded so a single-core machine still gets some batching benefit and a
    many-core machine doesn't hold an excessive number of docs in memory.
    """
    return max(16, min(_worker_process_count() * 16, 128))


# Rows per batch_analyzer.analyze_iterator() call. Batching lets spaCy's
# nlp.pipe() amortise per-call overhead across many texts instead of paying
# it once per row (the original bottleneck at scale), while still bounding
# how long cancel_check() can go unchecked and how coarse progress updates
# during the analysis phase get on very large files. Scales with worker
# count so a multi-process run keeps every worker fed with a full chunk.
ANALYSIS_CHUNK_SIZE = max(200, _analysis_batch_size() * _worker_process_count() * 4)


ENTITY_TYPE_OPTIONS = [
    ("Names (PERSON)", "PERSON"),
    ("Phone numbers", "PHONE_NUMBER"),
    ("Email addresses", "EMAIL_ADDRESS"),
    ("National Insurance numbers (UK_NINO)", "UK_NINO"),
    ("UK postcodes", "UK_POSTCODE"),
    ("Street addresses (UK_ADDRESS)", "UK_ADDRESS"),
    ("Housing / repair / tenancy references (HOUSING_REF)", "HOUSING_REF"),
    ("Key-safe / access codes (ACCESS_CODE)", "ACCESS_CODE"),
    ("Broader locations (LOCATION)", "LOCATION"),
    ("Dates / times (DATE_TIME)", "DATE_TIME"),
    ("Organisations (ORGANIZATION)", "ORGANIZATION"),
]

DEFAULT_ENTITY_TYPES = {
    "PERSON",
    "PHONE_NUMBER",
    "EMAIL_ADDRESS",
    "UK_NINO",
    "UK_POSTCODE",
    "UK_ADDRESS",
    "HOUSING_REF",
    "ACCESS_CODE",
}

ENTITY_SPECIFIC_LABELS = {
    "PERSON": "<PERSON>",
    "PHONE_NUMBER": "<PHONE_NUMBER>",
    "EMAIL_ADDRESS": "<EMAIL>",
    "UK_NINO": "<NINO>",
    "UK_POSTCODE": "<POSTCODE>",
    "UK_ADDRESS": "<ADDRESS>",
    "HOUSING_REF": "<HOUSING_REF>",
    "ACCESS_CODE": "<ACCESS_CODE>",
    "LOCATION": "<LOCATION>",
    "DATE_TIME": "<DATE_TIME>",
    "ORGANIZATION": "<ORGANIZATION>",
}


def default_selected_columns(columns: list[str]) -> list[str]:
    keywords = ["comment", "description", "note", "notes", "free", "text", "details", "message"]
    defaults = [
        column
        for column in columns
        if any(keyword in column.lower() for keyword in keywords)
    ]
    return defaults or columns[:1]


def anonymised_column_name(existing_columns: list[str], source_column: str) -> str:
    base_name = f"{source_column}_Anonymised"
    candidate = base_name
    suffix = 1
    while candidate in existing_columns:
        candidate = f"{base_name}_{suffix}"
        suffix += 1
    return candidate


def load_dataframe(file_path: str | Path) -> pd.DataFrame:
    path = Path(file_path)
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)
    return pd.read_excel(path)


def build_operators(selected_entity_types: Iterable[str], redaction_style: str) -> dict[str, OperatorConfig]:
    if redaction_style.startswith("Generic"):
        return {"DEFAULT": OperatorConfig("replace", {"new_value": "<REDACTED>"})}

    operators = {
        entity_type: OperatorConfig("replace", {"new_value": ENTITY_SPECIFIC_LABELS[entity_type]})
        for entity_type in selected_entity_types
        if entity_type in ENTITY_SPECIFIC_LABELS
    }
    operators["DEFAULT"] = OperatorConfig("replace", {"new_value": "<REDACTED>"})
    return operators


class AnonymisationCancelled(Exception):
    """Raised to unwind process_dataframe when the caller requests cancellation."""


def process_dataframe(
    df: pd.DataFrame,
    selected_columns: list[str],
    selected_entity_types: Iterable[str],
    redaction_style: str,
    score_threshold: float,
    progress_callback: Callable[[int, int, str], None] | None = None,
    progress_interval: int = 25,
    cancel_check: Callable[[], bool] | None = None,
) -> tuple[pd.DataFrame, list[dict], list[dict], dict[str, int | float]]:
    output_df = df.copy()
    analyzer = create_beyond_analyzer(score_threshold=score_threshold)
    anonymizer = AnonymizerEngine()
    operators = build_operators(selected_entity_types, redaction_style)
    selected_entity_types = set(selected_entity_types)
    scan_for_missed_names = "PERSON" in selected_entity_types

    results_summary: list[dict] = []
    residual_flags: list[dict] = []
    total_cells = len(output_df) * len(selected_columns)
    processed_cells = 0

    if progress_callback is not None:
        progress_callback(processed_cells, total_cells, "Starting anonymisation...")

    for source_column in selected_columns:
        output_column = anonymised_column_name(list(output_df.columns), source_column)
        texts = output_df[source_column].fillna("").astype(str).tolist()
        anonymised_texts: list[str] = []

        for chunk_start in range(0, len(texts), ANALYSIS_CHUNK_SIZE):
            if cancel_check is not None and cancel_check():
                raise AnonymisationCancelled("Anonymisation cancelled by user.")

            chunk_texts = texts[chunk_start : chunk_start + ANALYSIS_CHUNK_SIZE]
            # Calling nlp_engine.process_batch()/analyzer.analyze() directly
            # (rather than BatchAnalyzerEngine.analyze_iterator, which
            # discards them) keeps nlp_artifacts around per text - needed
            # below for scan_missed_proper_nouns, which reads spaCy's raw
            # POS tags and entity list rather than Presidio's filtered
            # output. Batching behaviour (and therefore throughput) is
            # unchanged: this is what analyze_iterator does internally.
            nlp_artifacts_batch = analyzer.nlp_engine.process_batch(
                texts=chunk_texts,
                language="en",
                batch_size=_analysis_batch_size(),
                n_process=_worker_process_count(),
            )

            for offset, (raw_text, nlp_artifacts) in enumerate(nlp_artifacts_batch):
                row_index = chunk_start + offset
                analysis = analyzer.analyze(
                    text=raw_text,
                    language="en",
                    nlp_artifacts=nlp_artifacts,
                    allow_list=PERSON_FALSE_POSITIVE_ALLOW_LIST,
                )
                redaction_targets = [
                    result for result in analysis if result.entity_type in selected_entity_types
                ]

                anonymised_result = anonymizer.anonymize(
                    text=raw_text,
                    analyzer_results=redaction_targets,
                    operators=operators,
                )
                anonymised_text = anonymised_result.text
                anonymised_texts.append(anonymised_text)

                for result in redaction_targets:
                    results_summary.append(
                        {
                            "source_column": source_column,
                            "output_column": output_column,
                            "row_index": row_index,
                            "entity_type": result.entity_type,
                            "score": round(result.score, 3),
                            "original_text": raw_text[result.start : result.end],
                            "start": result.start,
                            "end": result.end,
                        }
                    )

                residuals = residual_scan(anonymised_text)
                if scan_for_missed_names:
                    residuals = residuals + scan_missed_proper_nouns(
                        nlp_artifacts, redaction_targets, PERSON_FALSE_POSITIVE_ALLOW_LIST
                    )
                if residuals:
                    # A reviewer deciding "is this actually a name/address"
                    # previously had to leave the workbook and go find this
                    # exact row in the source file to see the surrounding
                    # sentence - the flagged text alone is often ambiguous
                    # out of context. Capturing a snippet here (while
                    # anonymised_text is in scope) means both the summary
                    # and per-row sheets can show it directly.
                    for finding in residuals:
                        ctx_start = max(0, finding["start"] - 40)
                        ctx_end = min(len(anonymised_text), finding["end"] + 40)
                        prefix = "..." if ctx_start > 0 else ""
                        suffix = "..." if ctx_end < len(anonymised_text) else ""
                        snippet = anonymised_text[ctx_start:ctx_end].replace("\n", " ").replace("\r", " ")
                        finding["context"] = f"{prefix}{snippet}{suffix}"
                    residual_flags.append(
                        {
                            "source_column": source_column,
                            "output_column": output_column,
                            "row_index": row_index,
                            "findings": residuals,
                        }
                    )

                processed_cells += 1
                if progress_callback is not None and (
                    processed_cells % progress_interval == 0 or processed_cells == total_cells
                ):
                    status_message = f"Processed {processed_cells} of {total_cells} cells..."
                    progress_callback(processed_cells, total_cells, status_message)

        output_df[output_column] = anonymised_texts

    stats = {
        "records_processed": len(output_df),
        "columns_anonymised": len(selected_columns),
        "pii_entities_detected": len(results_summary),
        "rows_with_possible_residual_items": len(residual_flags),
        "processed_cells": processed_cells,
        "total_cells": total_cells,
    }
    return output_df, results_summary, residual_flags, stats


def _excel_safe(value):
    """
    openpyxl's write-only Worksheet.append() writes each cell as it's
    called rather than building an in-memory grid, but it still rejects
    NaN (pandas' missing-value marker propagates into object columns too,
    not just numeric ones) since NaN has no valid Excel/XML representation.
    """
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def write_output_workbook(destination, df: pd.DataFrame, results_summary: list[dict], residual_flags: list[dict]) -> None:
    """
    Writes with openpyxl's write_only mode: rows are streamed to disk as
    they're appended instead of being held as Python Cell objects for the
    whole workbook. The standard pandas ExcelWriter/ws.to_excel() path
    builds the entire workbook in memory first, which measurably crashed
    with MemoryError on a Detection Report sheet with ~400k rows (a
    realistic size at ~100k input rows, since each cell can produce
    several detected entities) - this avoids that regardless of row count.
    """
    wb = Workbook(write_only=True)

    data_sheet = wb.create_sheet("Anonymised Data")
    data_sheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        data_sheet.append([_excel_safe(value) for value in row])

    detection_columns = [
        "source_column", "output_column", "row_index",
        "entity_type", "score", "original_text", "start", "end",
    ]
    detection_sheet = wb.create_sheet("Detection Report")
    detection_sheet.append(detection_columns)
    for item in results_summary:
        detection_sheet.append([_excel_safe(item.get(column)) for column in detection_columns])

    if residual_flags:
        # The same flagged value (e.g. a missed surname) typically recurs
        # across many rows in a large file - at 100k rows that's tens of
        # thousands of near-duplicate flag rows, which nobody actually
        # reviews row by row. This groups by (label, text) so "Munday"
        # appearing 200 times is one reviewable line, not 200, with a
        # sample of where to look if someone wants to check context.
        # Placed before the full per-row sheet, not instead of it.
        summary_sheet = wb.create_sheet("Residual Flags Summary")
        summary_sheet.append(["text", "label", "occurrences", "sample_context", "example_locations"])
        for text, label, count, sample_context, examples in _summarise_residual_flags(residual_flags):
            summary_sheet.append([
                _excel_safe(text),
                _excel_safe(label),
                _excel_safe(count),
                _excel_safe(sample_context),
                _excel_safe(examples),
            ])

        residual_columns = ["source_column", "output_column", "row_index", "label", "text", "context"]
        residual_sheet = wb.create_sheet("Residual Flags")
        residual_sheet.append(residual_columns)
        for item in residual_flags:
            for finding in item["findings"]:
                residual_sheet.append([
                    _excel_safe(item["source_column"]),
                    _excel_safe(item["output_column"]),
                    _excel_safe(item["row_index"]),
                    _excel_safe(finding["label"]),
                    _excel_safe(finding["text"]),
                    _excel_safe(finding.get("context", "")),
                ])

    wb.save(destination)


def _summarise_residual_flags(
    residual_flags: list[dict], max_examples: int = 3
) -> list[tuple[str, str, int, str, str]]:
    """
    Groups residual flags by (text, label) and returns one row per unique
    combination, most frequent first, with a small sample of where each
    one occurs. Frequency-first ordering surfaces the highest-volume
    items (most likely either a real recurring leak or common review-
    sheet noise) at the top rather than requiring a scroll through
    thousands of rows in file order to notice a pattern.

    Also carries one sample_context snippet per group - on data where most
    flagged values are unique (real customer names, unlike a synthetic test
    file's repeated small name pool), the frequency-based grouping barely
    reduces row count, so the context snippet is what actually saves a
    reviewer from opening the source file to judge each item: many
    "POSSIBLE_MISSED_NAME" flags are obviously real or obviously not once
    the surrounding sentence is visible, without needing to go find it.
    """
    groups: dict[tuple[str, str], dict] = {}
    for item in residual_flags:
        for finding in item["findings"]:
            key = (finding["text"], finding["label"])
            group = groups.setdefault(key, {"count": 0, "examples": [], "context": finding.get("context", "")})
            group["count"] += 1
            if len(group["examples"]) < max_examples:
                group["examples"].append(f"{item['source_column']} row {item['row_index']}")

    rows = []
    for (text, label), group in groups.items():
        examples = ", ".join(group["examples"])
        if group["count"] > len(group["examples"]):
            examples += f", +{group['count'] - len(group['examples'])} more"
        rows.append((text, label, group["count"], group["context"], examples))

    rows.sort(key=lambda row: row[2], reverse=True)
    return rows
